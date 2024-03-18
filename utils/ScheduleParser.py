import re
import copy

import requests
from collections import OrderedDict
import openpyxl
from lxml import etree
from lxml.etree import ParserError
from config import CS_URL, TIMETABLE_PATH

# TODO add logging
# TODO сериализация объекта
# TODO Распарсить ещё сами предметы
# TODO Через абстрактные классы преобразовывать расписание для дня, недели, месяца и т.д.
# TODO переделать дни, т.к они в словаре могут поменять порядок
# TODO Поиск преподавателей в определённый день
# TODO получать числитель/знаменатель через внутренний календарь


class ScheduleParser:
    """
    Класс для парсинга таблицы с расписанием занятий и преобразования ее в словарь объектов.
    """

    _tableObj: OrderedDict[any] = None
    _freeAudiences: OrderedDict[any] = None
    _audiences: set[str] = set()
    _calendar: None = None

    def __init__(self):
        self._toObject(self._parse(self._downloadTable()))
        self._makeFreeAudiences()

    def _downloadTable(self) -> str:
        try:
            parser = etree.HTMLParser()
            dom = etree.HTML(requests.get(CS_URL).content, parser)
        except ParserError as e:
            print(e)
        idTable = re.findall("\/d\/(.*?)\/", ",".join(dom.xpath("//a/@href")))[0]
        from utils.HTTPMethods import downloadFile
        downloadFile(f"https://docs.google.com/spreadsheets/d/{idTable}/export?format=xlsx&id={idTable}", TIMETABLE_PATH)
        return TIMETABLE_PATH

    def _parse(self, filename):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError as e:
            print(e)

        maxRows = sheet.max_row
        maxCols = sheet.max_column

        table_list = [[""] * maxCols for _ in range(maxRows)]

        for row in sheet.iter_rows():
            for cell in row:
                value = str(cell.value)
                r = cell.row - 1
                c = cell.column - 1
                table_list[r][c] = value

        for r in sheet.merged_cells.ranges:
            cl, rl, cr, rr = r.bounds
            value = str(r.start_cell.value)
            for c in range(cl - 1, cr):
                for row in range(rl - 1, rr):
                    table_list[row][c] = value

        delete = []
        for r in range(maxRows):
            remove = True
            for c in range(maxCols):
                if not table_list[r][c] in ["None", ""]:
                    remove = False
                    break
            if remove:
                delete.insert(0, r)
        for i in delete:
            maxRows -= 1
            table_list.pop(i)
        return table_list

    def _toObject(self, table):
        objects = OrderedDict()
        for i in range(2, len(table[0])):
            course = table[0][i].strip()
            group = table[1][i].strip()
            direction = table[2][i].strip()
            profile = table[3][i].strip()
            timetable = OrderedDict({"Числитель": OrderedDict(), "Знаменатель": OrderedDict()})
            numerator = True
            for d in range(4, len(table)):
                day = table[d][0].strip()
                time = table[d][1].strip().replace(" - ", "-").replace("-", " - ")
                subject = table[d][i].strip()

                regular = re.findall(
                    "(.*?) (преп\.|ст\.преп\.|доц\.|асс\.|проф\.) (.*?) (\d+.*?)$",
                    subject,
                )
                if len(regular) > 0:
                    _, rank, teacher, audience = regular[0]  # TODO subject
                    self._audiences.add(audience)
                regular = re.findall("(.*?) (.*?) (\d+\S)$", subject)
                if len(regular) > 0:
                    _, teacher, audience = regular[0]  # TODO subject
                    self._audiences.add(audience)
                key = "Числитель" if numerator else "Знаменатель"
                if not day in timetable[key]:
                    timetable[key][day] = OrderedDict()
                if not time in timetable[key][day]:
                    timetable[key][day][time] = subject
                numerator = not numerator

            if not course in objects:
                objects[course] = OrderedDict()
            if not direction in objects[course]:
                objects[course][direction] = OrderedDict()
            if not profile in objects[course][direction]:
                objects[course][direction][profile] = OrderedDict()
            if group in objects[course][direction][profile]:
                objects[course][direction][profile][
                    re.sub("(\d+)", lambda x: f"{x.group()}.1", group)
                ] = objects[course][direction][profile].pop(group)
                objects[course][direction][profile][
                    re.sub("(\d+)", lambda x: f"{x.group()}.2", group)
                ] = timetable
            else:
                objects[course][direction][profile][group] = timetable

        self._tableObj = objects
        return objects

    def _toText(self, object) -> str:
        return ""

    def getScheduleForDay(self, course, direction, profile, group, numerator, day):
        tempObj: OrderedDict[any] = copy.deepcopy(self._tableObj)
        daySchedule: str = f"{day} ({numerator})\n"

        for key in (course, direction, profile, group, numerator, day):
            if key in tempObj:
                tempObj = tempObj[key]
            else:
                if day == "Воскресенье":
                    # TODO переделать (запихнуть дни в расписание)
                    return daySchedule + "\nВыходной!"
                return ""

        # Проверка на выходной
        if (
            all(subject == "None" for subject in tempObj.values())
            or day == "Воскресенье"
        ):
            return daySchedule + "\nВыходной!"

        # Удаление лишних значений "None" в начале и в конце расписания
        while tempObj and tempObj[next(iter(tempObj))] == "None":
            tempObj.pop(next(iter(tempObj)))
        while tempObj and tempObj[next(reversed(tempObj))] == "None":
            tempObj.pop(next(reversed(tempObj)))

        # Формирование расписания
        for time, subject in tempObj.items():
            daySchedule += f'{time} => {subject.replace("None", "Окно")}\n\n'
        daySchedule = daySchedule[:-2]
        return daySchedule

    def getTableObj(self):
        if self._tableObj is None:
            self._toObject(self._parse(self._downloadTable()))
        return self._tableObj

    def getFreeAudiences(self, day, time, numerator):
        if len(self._freeAudiences) == 0:
            self._makeFreeAudiences()
        return "Вот список свободных аудиторий: " + ", ".join(
            self._freeAudiences[day][time][numerator]
        )

    def _makeFreeAudiences(self):  # TODO Сделать через таблицу
        tempObj: OrderedDict[any] = copy.deepcopy(self._tableObj)
        freeAudiences: OrderedDict[any] = OrderedDict()

        for course, directions in tempObj.items():
            for direction, profiles in directions.items():
                for profile, groups in profiles.items():
                    for group, numerators in groups.items():
                        for numerator, days in numerators.items():
                            for day, times in days.items():
                                for time, subject in times.items():
                                    if day in freeAudiences:
                                        if time in freeAudiences[day]:
                                            if numerator in freeAudiences[day][time]:
                                                audience = "-1"
                                                regular = re.findall(
                                                    "(.*?) (преп\.|ст\.преп\.|доц\.|асс\.|проф\.) (.*?) (\d+.*?)$",
                                                    subject,
                                                )
                                                if len(regular) > 0:
                                                    _, rank, teacher, audience = (
                                                        regular[0]
                                                    )
                                                    if (
                                                        audience
                                                        in freeAudiences[day][time][
                                                            numerator
                                                        ]
                                                    ):
                                                        freeAudiences[day][time][
                                                            numerator
                                                        ].remove(audience)
                                                    continue
                                                regular = re.findall(
                                                    "(.*?) (.*?) (\d+\S)$", subject
                                                )
                                                if len(regular) > 0:
                                                    _, teacher, audience = regular[0]
                                                    if (
                                                        audience
                                                        in freeAudiences[day][time][
                                                            numerator
                                                        ]
                                                    ):
                                                        freeAudiences[day][time][
                                                            numerator
                                                        ].remove(audience)
                                                    continue
                                            else:
                                                freeAudiences[day][time][numerator] = (
                                                    copy.copy(self._audiences)
                                                )
                                        else:
                                            freeAudiences[day][time] = OrderedDict()
                                    else:
                                        freeAudiences[day] = OrderedDict()
        self._freeAudiences = freeAudiences
        return freeAudiences


scheduleParser = ScheduleParser()
if __name__ == "__main__":
    table = scheduleParser.getScheduleForDay(
        "1 курс",
        'Направление "Прикладная информатика"',
        'профиль "Прикладная информатика в экономике"',
        "13 группа",
        "Числитель",
        "Пятница",
    )

    print(table)
    print(scheduleParser.getFreeAudiences("Понедельник", "8:00 - 9:35", "Числитель"))
