import asyncio
import copy
import datetime
import re
from collections import OrderedDict
from typing import Any

import aiohttp
import openpyxl
from config import TIMETABLE_PATH, TIMEZONE
from database.database import db
from database.models import Faculty, Group
from database.services.faculty import create_group_direction_profile_course, get_faculty_by_id
from database.services.group import bifurcation_group
from loguru import logger
from lxml import etree

# FIXME: bot-container       | sys:1: SAWarning: Multiple rows returned with uselist=False for eagerly-loaded attribute 'User.settings'
# TODO: Логотип в логах при запуске
# TODO: Подключить сюда redis
# TODO: Объединить выходные (объединить одинаковые дни)
# TODO: Оптимизировать получение расписания на неделю (ну очень долго)
# TODO: Property
# TODO: Фильтр фкновских аудиторий (мы чаще занимаемся на фкн, а пары в других факах появляются заметно реже)
# TODO: Поиск преподавателей в определённый день
# TODO: Сделать аннотации
# TODO: Логирование
# TODO: Сериализация объекта
# TODO: Через абстрактные классы преобразовывать расписание для дня, недели, месяца и т.д.
# TODO: Переделать дни, т.к они в словаре могут поменять порядок
# TODO: Получать числитель/знаменатель через внутренний календарь
# TODO: Правильная сортировка словаря с предметами
# TODO: Расписание для преподавателей (создаём расписание при создании общего)
# TODO: Получать расписание по объекту `User`
# TODO: Рефакторинг
# TODO: Жёсткий рефакторинг


class Subject:
    """
    Class to handle information about subject
    """

    def __init__(
        self,
        name: str,
        rank: str = None,
        teacher: str = None,
        audience: str = None,
        window: bool = False,
        time: str = None,
    ) -> None:
        self._name: str = name
        self._rank: str | None = rank
        self._teacher: str | None = teacher
        self._audience: str | None = audience
        self._window: bool = window
        self._time: str | None = time

    @property
    def is_window(self) -> bool:
        return self._window

    @property
    def audience(self) -> str | None:
        return self._audience

    @audience.setter
    def audience(self, audience) -> None:
        self._audience = audience

    @property
    def time(self) -> str | None:
        return self._time

    @time.setter
    def time(self, time) -> None:
        self._time = time

    def to_text(self, format_string: str = "name rank teacher audience") -> str:
        return (
            format_string.replace("name", self._name or "")
            .replace("rank", self._rank or "")
            .replace("teacher", self._teacher or "")
            .replace("audience", self._audience or "")
        )

    def to_dict(self) -> dict:
        return {
            "name": self._name,
            "rank": self._rank,
            "teacher": self._teacher,
            "audience": self._audience,
            "window": self._window,
            "time": self._time,
        }


class ScheduleParser:
    """
    Class to parse the timetable spreadsheet and convert it into a dictionary of objects.
    """

    _faculty_id: int = 1  #! TODO: Надо продумать этот момент
    _faculty: Faculty
    _timetable_url: str = None
    _time: set[str] = set()
    _table: list[list[str]] = None
    _tableObj: OrderedDict[Any] = None
    _freeAudiences: OrderedDict[Any] = None
    _audiences: set[str] = set()
    _temp_audiences: list[str] = list()

    async def __ainit__(self):
        async with db.session() as session:
            async with session.begin():
                self._faculty = await get_faculty_by_id(session, self._faculty_id)
        await self._toObject(await self._parse(await self._downloadTable()))
        await self._makeFreeAudiences()
        logger.success("Init schedule parser!")

    async def updateTable(self) -> None:
        """
        Update the timetable by downloading and parsing the latest data.
        """
        # TODO: Необходимо сделать уведомление о изменении связанное с парами
        # TODO: Изменённую пару добавить в APScheduler
        logger.info("Timetable begin update")
        self._time: set[str] = set()
        self._table: list[list[str]] = None
        self._tableObj: OrderedDict[Any] = None
        self._freeAudiences: OrderedDict[Any] = None
        self._audiences: set[str] = set()
        self._temp_audiences: list[str] = list()
        async with db.session() as session:
            async with session.begin():
                self._faculty = await get_faculty_by_id(session, self._faculty_id)
        await self._toObject(await self._parse(await self._downloadTable()))
        await self._makeFreeAudiences()
        logger.info("Timetable updated successfully")

    async def get_time(self) -> set[str]:
        return self._time

    @staticmethod
    def time_in_range(timerange: str, x: datetime.time):
        """Return true if x is in the timerange"""
        start, end = timerange.split(" - ")
        start = list(map(int, start.split(":")))
        start = datetime.time(hour=start[0], minute=start[1], tzinfo=TIMEZONE)
        end = list(map(int, end.split(":")))
        end = datetime.time(hour=end[0], minute=end[1], tzinfo=TIMEZONE)
        if start <= end:
            return start <= x <= end
        return start <= x or x <= end

    @staticmethod
    def time_from_times(timerange: str, times: set):
        answer = []
        start, end = timerange.split(" - ")
        start = list(map(int, start.split(":")))
        start = datetime.time(hour=start[0], minute=start[1], tzinfo=TIMEZONE)
        end = list(map(int, end.split(":")))
        end = datetime.time(hour=end[0], minute=end[1], tzinfo=TIMEZONE)
        for time in times:
            time_start, time_end = time.split(" - ")
            time_start = list(map(int, time_start.split(":")))
            time_start = datetime.time(hour=time_start[0], minute=time_start[1], tzinfo=TIMEZONE)
            time_end = list(map(int, time_end.split(":")))
            time_end = datetime.time(hour=time_end[0], minute=time_end[1], tzinfo=TIMEZONE)
            if start <= time_start <= end and start <= time_end <= end:
                answer.append(time)

        return answer

    @staticmethod
    def _trim_schedule(schedule: OrderedDict) -> OrderedDict:
        """Remove leading and trailing windows from the schedule."""
        # Remove leading windows
        while schedule and next(iter(schedule.values())).is_window:
            schedule.popitem(last=False)
        # Remove trailing windows
        while schedule and next(reversed(schedule.values())).is_window:
            schedule.popitem(last=True)
        return schedule

    async def _downloadTable(self) -> str:
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(self._faculty.timetable_url) as response:
                    if response.status == 200:
                        content = await response.read()

                        parser = etree.HTMLParser()
                        dom = etree.HTML(content, parser)

                        idTable = re.findall(r"\/d\/(.*?)\/", ",".join(dom.xpath("//a/@href")))[0]

                        # Асинхронно скачиваем файл
                        from utils.HTTP_methods import download_file

                        await download_file(
                            f"https://docs.google.com/spreadsheets/d/{idTable}/export?format=xlsx&id={idTable}",
                            TIMETABLE_PATH,
                        )

                        logger.opt(colors=True).debug("<g>Updated table</g>")
                        return TIMETABLE_PATH
                    else:
                        logger.error(f"Error downloading the timetable: HTTP status {response.status}")
        except (aiohttp.ClientError, etree.ParserError, IndexError) as e:
            logger.error(f"Error downloading the timetable: {e}")

        return ""

    async def _parse(self, filename):
        try:
            workbook = await asyncio.to_thread(openpyxl.load_workbook, filename)
            sheet = workbook.active
        except FileNotFoundError as e:
            logger.error(f"File not found: {e}")

        maxRows = sheet.max_row
        maxCols = sheet.max_column

        table_list = [[""] * maxCols for _ in range(maxRows)]

        for row in sheet.iter_rows():
            for cell in row:
                value = str(cell.value)
                r = cell.row - 1
                c = cell.column - 1
                table_list[r][c] = value

        for r in sheet.merged_cells.ranges:
            cl, rl, cr, rr = r.bounds
            value = str(r.start_cell.value)
            for c in range(cl - 1, cr):
                for row in range(rl - 1, rr):
                    table_list[row][c] = value

        self._table = [row for row in table_list if row and any(cell not in ["None", ""] for cell in row)]

        return self._table

    async def _toObject(self, table):
        objects = OrderedDict()
        prev_course = table[1][2].strip()
        for i in range(2, len(table[0])):
            course, group, direction, profile = (table[j][i].strip() for j in range(4))
            if course == "None":
                course = prev_course
            async with db.session() as session:
                async with session.begin():
                    now = datetime.date.today()
                    year = now.year - int(course.split()[0])
                    if datetime.date(year, 1, 1) > now > datetime.date(year, 8, 31):
                        year -= 1
                    _group = await create_group_direction_profile_course(
                        session, group, year, profile, direction, self._faculty.id, course
                    )
                async with session.begin():
                    if course in objects and direction in objects[course] and profile in objects[course][direction]:
                        if group in objects[course][direction][profile] and _group is not None:
                            await bifurcation_group(
                                session=session,
                                group_id=_group.id,
                                group_name=group,
                                profile_id=_group.profile_id,
                                year_of_study=_group.year_of_study,
                            )
            timetable = OrderedDict({"Числитель": OrderedDict(), "Знаменатель": OrderedDict()})
            numerator = True
            previous_subject: dict[Subject] = dict()
            for d in range(4, len(table)):
                day, time = table[d][0].strip(), table[d][1].strip().replace(" - ", "-").replace("-", " - ")
                subject = table[d][i].strip()

                self._time.add(time)
                if regular := re.findall(r"(.*?) (преп\.|ст\.преп\.|доц\.|асс\.|проф\.) (.*?) (\d+.*?)$", subject):
                    name, rank, teacher, audience = regular[0]  # TODO subject
                    subject_object = Subject(name=name, rank=rank, teacher=teacher, audience=audience)
                    self._audiences.add(audience)
                elif regular := re.findall(r"(.*?) (.*?) (\d+\S)$", subject):
                    name, teacher, audience = regular[0]  # TODO subject
                    subject_object = Subject(name=name, teacher=teacher, audience=audience)
                    self._audiences.add(audience)
                else:
                    subject_object = Subject(name="Окно" if subject == "None" else subject, window=subject == "None")

                subject_object.time = time
                key = "Числитель" if numerator else "Знаменатель"
                if day not in timetable[key]:
                    timetable[key][day] = OrderedDict()
                if time not in timetable[key][day]:
                    timetable[key][day][time] = subject_object
                    if (
                        key in previous_subject
                        and previous_subject[key]
                        and day in previous_subject[key]
                        and previous_subject[key][day]
                        and previous_subject[key][day].to_text() == subject_object.to_text()
                    ):
                        time1, time2 = previous_subject[key][day].time, subject_object.time
                        new_time = time1.split(" - ")[0] + " - " + time2.split(" - ")[1]
                        if time1 in timetable[key][day]:
                            timetable[key][day].pop(time1)
                        if time2 in timetable[key][day]:
                            timetable[key][day].pop(time2)

                        if new_time not in timetable[key][day]:
                            subject_object.time = new_time
                            timetable[key][day][new_time] = subject_object
                if key not in previous_subject:
                    previous_subject[key] = dict()
                if day not in previous_subject[key]:
                    previous_subject[key][day] = dict()
                previous_subject[key][day] = subject_object
                numerator = not numerator

            if course not in objects:
                objects[course] = OrderedDict()
            if direction not in objects[course]:
                objects[course][direction] = OrderedDict()
            if profile not in objects[course][direction]:
                objects[course][direction][profile] = OrderedDict()
            if group in objects[course][direction][profile]:
                objects[course][direction][profile][re.sub(r"(\d+)", lambda x: f"{x.group()}.1", group)] = objects[
                    course
                ][direction][profile].pop(group)
                objects[course][direction][profile][re.sub(r"(\d+)", lambda x: f"{x.group()}.2", group)] = timetable
            else:
                objects[course][direction][profile][group] = timetable

            prev_course = course

        self._tableObj = objects
        return objects

    async def getScheduleForTime(self, group, numerator, day, time) -> str | None:
        tempObj: OrderedDict[Any] = copy.deepcopy(self._tableObj)

        # Проверка на Воскресенье
        if day == "Воскресенье":
            return None

        profile = group.profile
        direction = profile.direction
        course = direction.course

        # Проходим по ключам, чтобы добраться до расписания
        for key in (course.name, direction.name, profile.name, group.name, numerator, day):
            if key in tempObj:
                tempObj = tempObj[key]
            else:
                return None

        # Проверка на выходной
        if all(subject.is_window for subject in tempObj.values()):
            return None

        # Удаление лишних значений "None" в начале и в конце расписания
        tempObj = self._trim_schedule(tempObj)

        # Приведение времени в нужный формат
        for time_tmp in tempObj.keys():
            if time == time_tmp.split(" - ")[0]:
                time = time_tmp

        # Получаем расписание на указанное время
        subject = tempObj.get(time, None)
        if subject is None:
            return None

        # Формирование результата
        subject_text = subject.to_text()

        # Возвращаем отформатированное расписание
        return f"{day} ({numerator})\n\n{time} => {subject_text}"

    async def getScheduleForDay(self, group: Group, numerator, day):
        tempObj: OrderedDict[Any] = copy.deepcopy(self._tableObj)
        daySchedule: str = f"{day} ({numerator})\n"

        profile = group.profile
        direction = profile.direction
        course = direction.course

        for key in (course.name, direction.name, profile.name, group.name, numerator, day):
            if key in tempObj:
                tempObj = tempObj[key]
            else:
                if day == "Воскресенье":
                    return daySchedule + "\nВыходной!"
                return ""

        # Проверка на выходной
        if all(subject.is_window for subject in tempObj.values()) or day == "Воскресенье":
            return daySchedule + "\nВыходной!"

        # Удаление лишних значений "None" в начале и в конце расписания
        tempObj = self._trim_schedule(tempObj)

        # Формирование расписания
        import locale

        locale.setlocale(locale.LC_ALL, "ru_RU.UTF-8")
        datetime_now = datetime.datetime.now(TIMEZONE)
        time_now = datetime_now.time()
        day_now = datetime_now.date().strftime("%A").capitalize()
        for time, subject in tempObj.items():
            now = self.time_in_range(time, time_now) if day_now == day else False
            daySchedule += f"{'<b><i><U>' if now else ''}{time} {'(Сейчас)' if now else ''} => {subject.to_text()}{'</U></i></b>' if now else ''}\n\n"

        return daySchedule[:-2]

    async def getTableObj(self):
        if self._tableObj is None:
            await self._toObject(await self._parse(await self._downloadTable()))

        return self.subjects_to_dict(copy.deepcopy(self._tableObj))

    @staticmethod
    def subjects_to_dict(tableObj: dict[Any]) -> dict[Any]:
        for course, directions in tableObj.items():
            for direction, profiles in directions.items():
                for profile, groups in profiles.items():
                    for group, numerators in groups.items():
                        for numerator, days in numerators.items():
                            for day, times in days.items():
                                for time, subject in times.items():
                                    tableObj[course][direction][profile][group][numerator][day][
                                        time
                                    ] = subject.to_dict()
        return tableObj

    async def getFreeAudiencesObj(self):
        if self._freeAudiences is None or len(self._freeAudiences) == 0:
            await self._makeFreeAudiences()
        return self._freeAudiences

    async def getFreeAudiences(self, day, time, numerator):
        if self._freeAudiences is None or len(self._freeAudiences) == 0:
            await self._makeFreeAudiences()
        return "Вот список свободных аудиторий: " + ", ".join(self._freeAudiences[day][time][numerator])

    def _filter_audiences(self):
        pass

    async def _makeFreeAudiences(self):
        tempObj: OrderedDict[Any] = copy.deepcopy(self._tableObj)

        freeAudiences: OrderedDict[Any] = OrderedDict()

        # TODO: адаптивный фильтр - что чаще встречается, то и добавляем

        for course, directions in tempObj.items():
            for direction, profiles in directions.items():
                for profile, groups in profiles.items():
                    for group, numerators in groups.items():
                        for numerator, days in numerators.items():
                            for day, times in days.items():
                                for timerange, subject in times.items():
                                    for time in self.time_from_times(timerange, self._time):
                                        if day in freeAudiences:
                                            if time in freeAudiences[day]:
                                                if numerator in freeAudiences[day][time]:
                                                    audience = "-1"
                                                    if isinstance(subject, Subject):
                                                        audience = subject.audience
                                                        if audience in freeAudiences[day][time][numerator]:
                                                            freeAudiences[day][time][numerator].remove(audience)
                                                        continue
                                                    else:
                                                        regular = re.findall(
                                                            r"(.*?) (преп\.|ст\.преп\.|доц\.|асс\.|проф\.) (.*?) (\d+.*?)$",
                                                            subject,
                                                        )
                                                        if len(regular) > 0:
                                                            _, rank, teacher, audience = regular[0]
                                                            if audience in freeAudiences[day][time][numerator]:
                                                                freeAudiences[day][time][numerator].remove(audience)
                                                            continue
                                                        regular = re.findall(r"(.*?) (.*?) (\d+\S)$", subject)
                                                        if len(regular) > 0:
                                                            _, teacher, audience = regular[0]
                                                            if audience in freeAudiences[day][time][numerator]:
                                                                freeAudiences[day][time][numerator].remove(audience)
                                                            continue
                                                else:
                                                    freeAudiences[day][time][numerator] = list(
                                                        copy.copy(self._audiences)
                                                    )
                                            else:
                                                freeAudiences[day][time] = OrderedDict()
                                        else:
                                            freeAudiences[day] = OrderedDict()
        self._freeAudiences = freeAudiences
        return freeAudiences


async def init_schedule_parser(scheduleParser):
    await scheduleParser.__ainit__()
    return scheduleParser


schedule_parser = ScheduleParser()
